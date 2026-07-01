from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd

from ..profiling.csv_utils import detect_csv_encoding, infer_csv_dialect, read_csv_auto
from ..profiling.stats import infer_excel_sheet_layout
from .base import BaseParser, ParsedFile


class TableParser(BaseParser):
    supported_suffixes = (".csv", ".xlsx", ".xls")
    kind = "table"

    def __init__(self, preview_rows: int, *, sample_rows: int | None = 20000) -> None:
        self.preview_rows = preview_rows
        self.sample_rows = sample_rows

    def _read_preview(self, path: Path) -> pd.DataFrame:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return read_csv_auto(path, nrows=max(self.preview_rows, 1))
        return pd.read_excel(path, nrows=max(self.preview_rows, 1))

    def _excel_sheet_names(self, path: Path) -> list[str]:
        if path.suffix.lower() not in {".xlsx", ".xls"}:
            return []
        try:
            xls = pd.ExcelFile(path)
            try:
                return [str(x) for x in xls.sheet_names if str(x).strip()]
            finally:
                xls.close()
        except Exception:
            return []

    def _records(self, df: pd.DataFrame) -> list[dict[str, Any]]:
        preview = []
        for rec in df.head(self.preview_rows).to_dict(orient="records"):
            fixed = {}
            for k, v in rec.items():
                fixed[str(k)] = self._json_safe(v)
            preview.append(fixed)
        return preview

    def _raw_rows(self, df: pd.DataFrame) -> list[list[Any]]:
        """Preserve the literal top-left Excel cells before pandas picks a header row."""
        rows: list[list[Any]] = []
        for values in df.head(self.preview_rows).itertuples(index=False, name=None):
            rows.append([self._json_safe(v) for v in values])
        return rows

    def _json_safe(self, value: Any) -> Any:
        try:
            if isinstance(value, pd.Timestamp):
                return value.isoformat()
            if isinstance(value, pd.Timedelta):
                return str(value)
            if pd.isna(value):
                return None
            if hasattr(value, "item"):
                return value.item()
            return value
        except Exception:
            return str(value)

    def _estimate_rows(self, path: Path, preview_len: int) -> tuple[int, bool]:
        if path.suffix.lower() != ".csv":
            return preview_len, False
        try:
            # 快速按换行估计 CSV 行数；比 pandas 全量读入轻很多。
            with path.open("rb") as f:
                rows = sum(chunk.count(b"\n") for chunk in iter(lambda: f.read(1024 * 1024), b""))
            return max(0, rows - 1), True
        except Exception:
            return preview_len, False

    def _excel_shapes(self, path: Path) -> dict[str, list[int]]:
        if path.suffix.lower() != ".xlsx":
            return {}
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            shapes: dict[str, list[int]] = {}
            for ws in wb.worksheets:
                # max_row includes the header row for ordinary tabular sheets.
                rows = max(0, int(ws.max_row or 0) - 1)
                cols = max(0, int(ws.max_column or 0))
                shapes[str(ws.title)] = [rows, cols]
            wb.close()
            return shapes
        except Exception:
            return {}

    def _parse_excel(self, path: Path) -> ParsedFile:
        xls = pd.ExcelFile(path)
        try:
            sheet_names = [str(x) for x in xls.sheet_names if str(x).strip()]
            shapes = self._excel_shapes(path)
            sheets: list[dict[str, Any]] = []
            primary_df = pd.DataFrame()
            primary_sheet = sheet_names[0] if sheet_names else ""
            union_columns: list[str] = []
            for sheet in sheet_names:
                raw_preview: list[list[Any]] = []
                raw_preview_rows_used = 0
                raw_preview_error = ""
                try:
                    raw_df = xls.parse(sheet_name=sheet, header=None, nrows=max(self.preview_rows, 1))
                    raw_preview = self._raw_rows(raw_df)
                    raw_preview_rows_used = int(len(raw_df))
                except Exception as exc:  # noqa: BLE001
                    raw_preview_error = str(exc)[:500]
                try:
                    df = xls.parse(sheet_name=sheet, nrows=max(self.preview_rows, 1))
                except Exception as exc:  # noqa: BLE001
                    item = {
                        "sheet_name": sheet,
                        "error": str(exc)[:500],
                        "raw_preview": raw_preview,
                        "raw_preview_rows_used": raw_preview_rows_used,
                    }
                    if raw_preview_error:
                        item["raw_preview_error"] = raw_preview_error
                    sheets.append(item)
                    continue
                if primary_df.empty:
                    primary_df = df
                    primary_sheet = sheet
                shape = shapes.get(sheet)
                shape_estimated = False
                if not shape:
                    shape = [int(len(df)), int(df.shape[1])]
                    shape_estimated = True
                columns = [str(c) for c in df.columns.tolist()]
                layout = infer_excel_sheet_layout(
                    raw_preview=raw_preview,
                    default_columns=columns,
                    sheet_name=sheet,
                    shape=shape,
                )
                sheets.append(
                    {
                        "sheet_name": sheet,
                        "shape": shape,
                        "shape_estimated": shape_estimated,
                        "preview_rows_used": int(len(df)),
                        "raw_preview_rows_used": raw_preview_rows_used,
                        "columns": columns,
                        "dtypes": {str(k): str(v) for k, v in df.dtypes.to_dict().items()},
                        "preview": self._records(df),
                        "raw_preview": raw_preview,
                        **layout,
                        **({"raw_preview_error": raw_preview_error} if raw_preview_error else {}),
                    }
                )
                for col in columns:
                    if col not in union_columns:
                        union_columns.append(col)
            if primary_df.empty and sheet_names:
                primary_df = xls.parse(sheet_name=sheet_names[0], nrows=max(self.preview_rows, 1))
            primary_columns = [str(c) for c in primary_df.columns.tolist()]
            columns = union_columns or primary_columns
            shape = shapes.get(primary_sheet) or [int(len(primary_df)), int(primary_df.shape[1])]
            summary = (
                f"Excel 工作簿: {len(sheet_names)} 个 sheet；已逐 sheet 读取表头/shape/前 {self.preview_rows} 行原始切片；"
                f"默认主 sheet `{primary_sheet}` 行列: {shape[0]} x {shape[1]}"
            )
            return ParsedFile(
                path=path,
                kind=self.kind,
                text_summary=summary,
                preview=self._records(primary_df),
                columns=columns,
                metadata={
                    "shape": shape,
                    "shape_estimated": primary_sheet not in shapes,
                    "preview_rows_used": int(len(primary_df)),
                    "dtypes": {str(k): str(v) for k, v in primary_df.dtypes.to_dict().items()},
                    "excel_sheet_names": sheet_names,
                    "excel_default_sheet": primary_sheet,
                    "excel_primary_columns": primary_columns,
                    "excel_sheets": sheets,
                    "excel_read_policy": (
                        "Every sheet is inspected with shape/header/default preview and header=None raw top rows; "
                        "downstream code should specify sheet_name explicitly for important sheets."
                    ),
                    "source_format": path.suffix.lower().lstrip("."),
                },
            )
        finally:
            xls.close()

    def parse(self, path: Path) -> ParsedFile:
        if path.suffix.lower() in {".xlsx", ".xls"}:
            return self._parse_excel(path)
        df = self._read_preview(path)
        excel_sheet_names = self._excel_sheet_names(path)
        row_count, estimated = self._estimate_rows(path, len(df))
        preview_df = df.head(self.preview_rows).copy()
        preview = self._records(preview_df)
        summary = f"表格行列: {row_count} x {df.shape[1]}"
        if estimated:
            summary += "（CSV 行数快速估计，字段类型基于预览切片）"
        dialect = infer_csv_dialect(path) if path.suffix.lower() == ".csv" else None
        return ParsedFile(
            path=path,
            kind=self.kind,
            text_summary=summary,
            preview=preview,
            columns=[str(c) for c in df.columns.tolist()],
            metadata={
                "shape": [int(row_count), int(df.shape[1])],
                "shape_estimated": bool(estimated),
                "preview_rows_used": int(len(df)),
                "dtypes": {str(k): str(v) for k, v in df.dtypes.to_dict().items()},
                "csv_dialect": dialect.__dict__ if dialect else None,
                "csv_encoding": detect_csv_encoding(path) if path.suffix.lower() == ".csv" else "",
                "excel_sheet_names": excel_sheet_names,
                "excel_default_sheet": excel_sheet_names[0] if excel_sheet_names else "",
                "source_format": path.suffix.lower().lstrip("."),
            },
        )
