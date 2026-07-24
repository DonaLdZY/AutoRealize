from __future__ import annotations

from dataclasses import dataclass, field
import json
from pathlib import Path
import re
from typing import Any
import warnings

import numpy as np
import pandas as pd

from .csv_utils import read_csv_auto
from ..utils.json_table import read_json_as_table


@dataclass
class ColumnProfile:
    name: str
    dtype: str
    null_ratio: float
    unique_count: int
    row_count: int = 0
    null_count: int = 0
    non_null_count: int = 0
    logical_type: str = "unknown"
    numeric_parse_ratio: float = 0.0
    datetime_parse_ratio: float = 0.0
    format_hints: list[str] = field(default_factory=list)
    sample_values: list[Any] = field(default_factory=list)
    top_values: list[str] = field(default_factory=list)
    value_pattern_hints: list[str] = field(default_factory=list)
    numeric_stats: dict[str, float | int | bool] = field(default_factory=dict)
    abnormal_tokens: list[str] = field(default_factory=list)
    quantiles: dict[str, float] = field(default_factory=dict)
    datetime_stats: dict[str, str] = field(default_factory=dict)


def read_table(
    path: Path,
    *,
    json_flatten_sep: str = "__",
    json_flatten_max_level: int | None = None,
    json_keep_raw_nested_columns: bool = False,
    max_rows: int | None = None,
    sheet_name: str | int | None = 0,
) -> pd.DataFrame:
    """读取可表格化文件。

    max_rows 用于数据认知预览的安全采样，避免数 GB CSV 在认知阶段被全量加载。
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_auto(path, nrows=max_rows)
    if suffix == ".json":
        df, _ = read_json_as_table(
            path,
            sep=json_flatten_sep,
            max_level=json_flatten_max_level,
            keep_raw_nested_columns=json_keep_raw_nested_columns,
        )
        return df.head(max_rows) if max_rows is not None else df
    return pd.read_excel(path, sheet_name=sheet_name, nrows=max_rows)


def table_probe_sample_rows(path: Path, *, configured_rows: int | None, large_threshold_bytes: int) -> int | None:
    """返回字段统计应读取的行数。None 表示允许全量读取。"""
    if configured_rows is None:
        return None
    try:
        # The current policy always applies the configured profiling cap, but we
        # still inspect the threshold so downstream metadata can distinguish
        # ordinary capped profiling from genuinely large-file protection.
        _ = path.stat().st_size > max(1, int(large_threshold_bytes))
    except Exception:
        pass
    return max(1, int(configured_rows))


def table_sampling_metadata(
    path: Path,
    *,
    configured_rows: int | None,
    large_threshold_bytes: int,
    rows_read: int | None = None,
) -> dict[str, Any]:
    """Describe the deterministic profiling sampling policy for a table."""
    try:
        file_size = int(path.stat().st_size)
    except Exception:
        file_size = 0
    threshold = max(1, int(large_threshold_bytes))
    max_rows = table_probe_sample_rows(
        path,
        configured_rows=configured_rows,
        large_threshold_bytes=large_threshold_bytes,
    )
    is_large_file = bool(file_size and file_size > threshold)
    if max_rows is None:
        reason = "full_scan_allowed"
    elif is_large_file:
        reason = "large_file_row_cap"
    else:
        reason = "configured_row_cap"
    return {
        "configured_max_rows": max_rows,
        "rows_read": rows_read,
        "file_size_bytes": file_size,
        "large_threshold_bytes": threshold,
        "is_large_file": is_large_file,
        "sampling_reason": reason,
        "sampled": bool(max_rows is not None),
    }


def profile_excel_sheets(
    path: Path,
    *,
    max_rows: int | None,
    top_k: int = 10,
    max_profile_columns: int = 80,
    preview_rows: int = 10,
    large_threshold_bytes: int = 256 * 1024 * 1024,
    full_profile_sheet_threshold: int = 10,
    representatives_per_group: int = 1,
) -> list[dict[str, Any]]:
    """Return compact per-sheet profiles for Excel workbooks.

    Policy:
    - Every sheet gets a lightweight inventory: name, shape/header, dtypes and a
      small preview.
    - If a workbook has only a few sheets and is not large, every sheet is
      deeply profiled with a full read.
    - Otherwise sheets with the same/near-same headers are grouped, and only
      representatives are deeply profiled. Non-representatives keep lightweight
      inventory so downstream agents still know the full workbook boundary.
    """
    if path.suffix.lower() not in {".xlsx", ".xls"}:
        return []
    try:
        xls = pd.ExcelFile(path)
    except Exception:
        return []
    try:
        sheet_names = [str(x) for x in xls.sheet_names if str(x).strip()]
        sheet_count = len(sheet_names)
        try:
            file_size = int(path.stat().st_size)
        except Exception:
            file_size = 0
        is_large = bool(file_size and file_size > max(1, int(large_threshold_bytes)))
        full_profile_all = sheet_count <= max(1, int(full_profile_sheet_threshold)) and not is_large
        shapes = _excel_sheet_shapes(path)

        inventories: list[dict[str, Any]] = []
        for sheet in sheet_names:
            raw_preview: list[list[Any]] = []
            raw_preview_rows_used = 0
            raw_preview_error = ""
            try:
                raw_df = xls.parse(sheet_name=sheet, header=None, nrows=max(1, int(preview_rows)))
                raw_preview = _raw_rows(raw_df.head(max(1, int(preview_rows))))
                raw_preview_rows_used = int(len(raw_df))
            except Exception as exc:  # noqa: BLE001
                raw_preview_error = str(exc)[:500]
            try:
                preview_df = xls.parse(sheet_name=sheet, nrows=max(1, int(preview_rows)))
            except Exception as exc:  # noqa: BLE001
                item = {
                    "sheet_name": sheet,
                    "error": str(exc)[:500],
                    "raw_preview": raw_preview,
                    "raw_preview_rows_used": raw_preview_rows_used,
                }
                if raw_preview_error:
                    item["raw_preview_error"] = raw_preview_error
                inventories.append(item)
                continue
            shape = shapes.get(sheet)
            shape_estimated = False
            if not shape:
                shape = [int(preview_df.shape[0]), int(preview_df.shape[1])]
                shape_estimated = True
            columns = [str(c) for c in preview_df.columns.tolist()]
            layout = infer_excel_sheet_layout(
                raw_preview=raw_preview,
                default_columns=columns,
                sheet_name=sheet,
                shape=shape,
            )
            inventories.append(
                {
                    "sheet_name": sheet,
                    "shape": shape,
                    "shape_estimated": shape_estimated,
                    "preview_rows_used": int(len(preview_df)),
                    "columns": columns,
                    "dtypes": {str(k): str(v) for k, v in preview_df.dtypes.to_dict().items()},
                    "preview": _records(preview_df.head(max(1, int(preview_rows)))),
                    "raw_preview": raw_preview,
                    "raw_preview_rows_used": raw_preview_rows_used,
                    **({"raw_preview_error": raw_preview_error} if raw_preview_error else {}),
                    "sheet_name_pattern": _normalize_sheet_name(sheet),
                    "header_signature": _header_signature(columns),
                    **layout,
                }
            )

        groups = _group_excel_sheet_inventories(inventories)
        representative_sheets: set[str] = set()
        for group in groups:
            reps = group["sheets"][: max(1, int(representatives_per_group))]
            representative_sheets.update(str(x) for x in reps)
        if sheet_names:
            representative_sheets.add(sheet_names[0])

        out: list[dict[str, Any]] = []
        for inv in inventories:
            sheet = str(inv.get("sheet_name", "") or "")
            if inv.get("error"):
                out.append(inv)
                continue
            group = next((g for g in groups if sheet in g.get("sheets", [])), {})
            is_representative = sheet in representative_sheets
            should_deep_profile = bool(full_profile_all or is_representative)
            profile_rows_limit = None if full_profile_all else max_rows
            entry = {
                **inv,
                "workbook_sheet_count": sheet_count,
                "workbook_file_size_bytes": file_size,
                "workbook_is_large": is_large,
                "profile_policy": "full_all_sheets" if full_profile_all else "representative_by_sheet_group",
                "sheet_group_id": group.get("group_id", ""),
                "sheet_group_size": group.get("sheet_count", 1),
                "sheet_group_sheets": group.get("sheets", [sheet])[:20],
                "sheet_group_representative": group.get("representative", sheet),
                "is_deep_profiled": should_deep_profile,
                "profile_rows_limit": profile_rows_limit,
            }
            if not should_deep_profile:
                out.append(entry)
                continue
            try:
                df = xls.parse(sheet_name=sheet, nrows=profile_rows_limit)
            except Exception as exc:  # noqa: BLE001
                entry["profile_error"] = str(exc)[:500]
                out.append(entry)
                continue
            use_cols = list(df.columns[:max_profile_columns])
            profiles = profile_dataframe(df[use_cols], top_k=top_k) if use_cols else []
            entry.update(
                {
                    "shape_profiled": [int(df.shape[0]), int(df.shape[1])],
                    "shape_sampled": [int(df.shape[0]), int(df.shape[1])],
                    "profiled_column_count": len(use_cols),
                    "column_profiles": [column_profile_to_dict(p) for p in profiles],
                }
            )
            out.append(entry)
        return out
    finally:
        xls.close()


def excel_sheet_groups_from_profiles(sheet_profiles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Return workbook-level sheet groups from per-sheet profiles."""
    return _group_excel_sheet_inventories(sheet_profiles)


def infer_excel_sheet_layout(
    *,
    raw_preview: list[list[Any]] | None,
    default_columns: list[str] | None,
    sheet_name: str = "",
    shape: list[int] | tuple[int, int] | None = None,
) -> dict[str, Any]:
    """Infer a conservative Excel sheet read strategy from top-left cells.

    The goal is not to fully understand the sheet. It is to prevent downstream
    code from blindly trusting pandas' default header when the opening rows look
    like raw data, notes, key-value documentation, or a non-zero header row.
    """

    rows = raw_preview if isinstance(raw_preview, list) else []
    rows = [row for row in rows if isinstance(row, list)]
    columns = [str(c) for c in (default_columns or [])]
    ncols = max([len(row) for row in rows] + [len(columns), 0])
    if shape and len(shape) >= 2:
        try:
            ncols = max(ncols, int(shape[1]))
        except Exception:
            pass
    if not rows and not columns:
        return {
            "layout_kind": "empty_or_unreadable",
            "read_strategy_kind": "inspect_manually",
            "header_confidence": 0.0,
            "detected_header_row": None,
            "recommended_read": _excel_read_example(sheet_name, header="inspect"),
            "reading_risks": ["Sheet could not be previewed; inspect manually before modeling."],
        }

    row_infos = [_row_layout_info(row, ncols=ncols) for row in rows[:12]]
    non_empty_rows = [info for info in row_infos if info["non_empty"] > 0]
    dense_rows = [info for info in row_infos if info["density"] >= 0.5 and info["non_empty"] >= 2]
    sparse_ratio = 1.0 - (sum(info["non_empty"] for info in row_infos) / max(1, len(row_infos) * max(1, ncols)))
    long_or_note_rows = sum(1 for info in non_empty_rows if info["long_text_count"] > 0 or info["note_marker_count"] > 0)
    one_or_two_cell_rows = sum(1 for info in non_empty_rows if info["non_empty"] <= 2)
    default_suspicious = _default_columns_look_suspicious(columns)

    best_idx: int | None = None
    best_score = -999.0
    for idx, info in enumerate(row_infos[:8]):
        score = _header_candidate_score(info)
        if idx + 1 < len(row_infos):
            score -= 1.25 * _type_pattern_similarity(info["type_pattern"], row_infos[idx + 1]["type_pattern"])
        if idx + 2 < len(row_infos):
            score -= 0.50 * _type_pattern_similarity(info["type_pattern"], row_infos[idx + 2]["type_pattern"])
        if info["note_marker_count"]:
            score -= 0.75
        if score > best_score:
            best_idx = idx
            best_score = score

    first_score = _header_candidate_score(row_infos[0]) if row_infos else -999.0
    best_conf = round(max(0.0, min(0.99, (best_score - 1.0) / 4.5)), 3)
    risks: list[str] = []
    if default_suspicious:
        risks.append("Pandas default columns look suspicious; verify header handling before modeling.")
    if rows:
        risks.append("Use header=None preview when validating sheet layout; opening rows may contain notes or raw data.")
    first_row_cells = [str(x).strip() for x in (rows[0] if rows else []) if _cell_text(x)]
    default_column_cells = [str(x).strip() for x in columns if str(x).strip()]
    default_columns_equal_first_row = bool(
        first_row_cells
        and default_column_cells
        and default_column_cells[: len(first_row_cells)] == first_row_cells[: len(default_column_cells)]
    )
    downstream_pattern_similarity = (
        max(
            (_type_pattern_similarity(row_infos[0]["type_pattern"], info["type_pattern"]) for info in row_infos[1:4]),
            default=0.0,
        )
        if row_infos
        else 0.0
    )
    short_code_like_first_row = bool(
        first_row_cells
        and len(first_row_cells) <= 4
        and all(_looks_like_identifier_value(value) for value in first_row_cells)
        and downstream_pattern_similarity >= 0.75
    )

    layout_kind = "standard_table"
    detected_header_row: int | None = 0
    read_strategy_kind = "default_header"

    best_info = row_infos[best_idx] if best_idx is not None and 0 <= best_idx < len(row_infos) else {}
    best_is_strong_header = (
        best_idx is not None
        and best_idx > 0
        and best_score >= 2.4
        and float(best_info.get("text_ratio") or 0.0) >= 0.75
        and float(best_info.get("numeric_ratio") or 0.0) <= 0.25
        and int(best_info.get("note_marker_count") or 0) == 0
    )

    if not non_empty_rows:
        layout_kind = "empty_or_unreadable"
        detected_header_row = None
        read_strategy_kind = "inspect_manually"
    elif best_is_strong_header:
        layout_kind = "non_default_header"
        detected_header_row = int(best_idx)
        read_strategy_kind = "explicit_header_row"
        risks.append(f"Likely header row is {best_idx}, not the first row.")
    elif (
        long_or_note_rows >= max(1, len(non_empty_rows) // 2)
        and one_or_two_cell_rows >= max(1, int(len(non_empty_rows) * 0.6))
    ):
        layout_kind = "document_like_sheet"
        detected_header_row = None
        read_strategy_kind = "header_none_document"
        risks.append("Sheet looks like notes/rules/key-value text, not an ordinary dataframe.")
    elif sparse_ratio >= 0.72 and len(dense_rows) <= 1:
        layout_kind = "sparse_or_irregular_sheet"
        detected_header_row = None
        read_strategy_kind = "header_none_inspect"
        risks.append("Sheet is sparse or irregular; do not assume rectangular tabular semantics.")
    elif (
        default_suspicious
        or first_score < 2.2
        or _first_row_matches_data_pattern(row_infos)
        or (default_columns_equal_first_row and short_code_like_first_row)
    ):
        layout_kind = "headerless_table"
        detected_header_row = None
        read_strategy_kind = "header_none_table"
        risks.append("First row looks like data rather than field names; use header=None or assign columns explicitly.")
        if default_columns_equal_first_row and short_code_like_first_row:
            risks.append("Default pandas columns repeat the first raw row and look like identifier values; the first data row would be lost with header=0.")
    else:
        detected_header_row = 0
        read_strategy_kind = "default_header"

    if layout_kind == "standard_table":
        risks = [risk for risk in risks if "suspicious" not in risk and "header=None preview" not in risk][:3]

    return {
        "layout_kind": layout_kind,
        "read_strategy_kind": read_strategy_kind,
        "header_confidence": best_conf if layout_kind in {"standard_table", "non_default_header"} else round(1.0 - best_conf, 3),
        "detected_header_row": detected_header_row,
        "recommended_read": _excel_read_example(sheet_name, header=detected_header_row if detected_header_row is not None else None, layout_kind=layout_kind),
        "reading_risks": list(dict.fromkeys(risks))[:6],
    }


def _excel_read_example(sheet_name: str, *, header: int | str | None = 0, layout_kind: str = "") -> str:
    sheet_part = f", sheet_name={sheet_name!r}" if sheet_name else ", sheet_name=<sheet_name>"
    if header == "inspect":
        return f"pd.read_excel(path{sheet_part}, header=None)  # inspect layout first"
    if header is None:
        suffix = "  # document-like sheet" if layout_kind == "document_like_sheet" else ""
        return f"pd.read_excel(path{sheet_part}, header=None){suffix}"
    if int(header) == 0:
        return f"pd.read_excel(path{sheet_part})"
    return f"pd.read_excel(path{sheet_part}, header={int(header)})"


def _row_layout_info(row: list[Any], *, ncols: int) -> dict[str, Any]:
    cells = [_cell_text(v) for v in row]
    if len(cells) < ncols:
        cells.extend([""] * (ncols - len(cells)))
    non_empty = [c for c in cells if c]
    type_pattern = [_cell_kind(c) for c in cells]
    text_count = sum(1 for c in non_empty if _cell_kind(c) == "text")
    numeric_count = sum(1 for c in non_empty if _cell_kind(c) == "number")
    long_text_count = sum(1 for c in non_empty if len(c) >= 50)
    note_marker_count = sum(1 for c in non_empty if _looks_like_note_marker(c))
    avg_len = sum(len(c) for c in non_empty) / max(1, len(non_empty))
    return {
        "non_empty": len(non_empty),
        "density": len(non_empty) / max(1, ncols),
        "text_ratio": text_count / max(1, len(non_empty)),
        "numeric_ratio": numeric_count / max(1, len(non_empty)),
        "unique_ratio": len(set(non_empty)) / max(1, len(non_empty)),
        "long_text_count": long_text_count,
        "note_marker_count": note_marker_count,
        "avg_len": avg_len,
        "type_pattern": type_pattern,
    }


def _header_candidate_score(info: dict[str, Any]) -> float:
    non_empty = int(info.get("non_empty") or 0)
    if non_empty < 2:
        return -4.0
    score = 0.0
    score += float(info.get("density") or 0.0) * 1.8
    score += float(info.get("text_ratio") or 0.0) * 2.2
    score += float(info.get("unique_ratio") or 0.0) * 1.2
    score -= float(info.get("numeric_ratio") or 0.0) * 1.7
    score -= min(2.0, float(info.get("long_text_count") or 0.0) * 0.8)
    score -= min(2.0, float(info.get("note_marker_count") or 0.0) * 0.9)
    if float(info.get("avg_len") or 0.0) > 32:
        score -= 0.8
    return score


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return ""
    return re.sub(r"\s+", " ", text)


def _cell_kind(text: str) -> str:
    if not text:
        return "empty"
    if re.fullmatch(r"[-+]?(?:\d+\.\d+|\d+|\.\d+)(?:%?)", text):
        return "number"
    if re.search(r"\d{4}[-/年]\d{1,2}", text):
        return "date"
    return "text"


def _looks_like_note_marker(text: str) -> bool:
    low = text.lower().strip()
    markers = [
        "note",
        "notes",
        "remark",
        "remarks",
        "description",
        "instruction",
        "instructions",
        "readme",
        "说明",
        "备注",
        "注释",
        "规则",
        "口径",
        "注意",
    ]
    return any(marker in low for marker in markers)


def _looks_like_identifier_value(text: str) -> bool:
    value = str(text or "").strip()
    if not value or " " in value or len(value) > 36:
        return False
    if _cell_kind(value) in {"number", "date"}:
        return True
    # Compact code values usually mix letters/digits/separators and contain no
    # natural-language header suffix. This remains evidence, not a hard schema.
    header_suffixes = ("代码", "名称", "编号", "时间", "日期", "数量", "类型", "地址", "说明", "规则")
    if any(value.endswith(suffix) for suffix in header_suffixes):
        return False
    has_letter = bool(re.search(r"[A-Za-z]", value))
    has_digit = bool(re.search(r"\d", value))
    return (has_letter and has_digit) or bool(re.fullmatch(r"[A-Z]{2,}[A-Z0-9_-]*", value))


def _default_columns_look_suspicious(columns: list[str]) -> bool:
    if not columns:
        return True
    cleaned = [str(c).strip() for c in columns if str(c).strip()]
    if not cleaned:
        return True
    unnamed = sum(1 for c in cleaned if c.lower().startswith("unnamed"))
    numeric = sum(1 for c in cleaned if _cell_kind(c) in {"number", "date"})
    long_text = sum(1 for c in cleaned if len(c) >= 50)
    return (
        unnamed / max(1, len(cleaned)) >= 0.25
        or numeric / max(1, len(cleaned)) >= 0.5
        or long_text / max(1, len(cleaned)) >= 0.25
    )


def _type_pattern_similarity(a: list[str], b: list[str]) -> float:
    length = max(len(a), len(b), 1)
    aa = list(a) + ["empty"] * (length - len(a))
    bb = list(b) + ["empty"] * (length - len(b))
    comparable = [(x, y) for x, y in zip(aa, bb) if x != "empty" or y != "empty"]
    if not comparable:
        return 0.0
    return sum(1 for x, y in comparable if x == y) / max(1, len(comparable))


def _first_row_matches_data_pattern(row_infos: list[dict[str, Any]]) -> bool:
    if len(row_infos) < 2:
        return False
    first = row_infos[0]
    if int(first.get("non_empty") or 0) < 2:
        return False
    sim = _type_pattern_similarity(first["type_pattern"], row_infos[1]["type_pattern"])
    if len(row_infos) >= 3:
        sim = max(sim, _type_pattern_similarity(first["type_pattern"], row_infos[2]["type_pattern"]))
    return sim >= 0.82 and float(first.get("numeric_ratio") or 0.0) >= 0.25


def _records(df: pd.DataFrame) -> list[dict[str, Any]]:
    return [
        {str(k): _json_safe(v) for k, v in row.items()}
        for row in df.to_dict(orient="records")
    ]


def _raw_rows(df: pd.DataFrame) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for values in df.itertuples(index=False, name=None):
        rows.append([_json_safe(v) for v in values])
    return rows


def _json_safe(value: Any) -> Any:
    try:
        if isinstance(value, pd.Timestamp):
            return value.isoformat()
        if isinstance(value, pd.Timedelta):
            return str(value)
        if pd.isna(value):
            return None
        if hasattr(value, "item"):
            return value.item()
    except Exception:
        pass
    return value


def _excel_sheet_shapes(path: Path) -> dict[str, list[int]]:
    """Best-effort sheet shapes without full pandas reads."""
    if path.suffix.lower() != ".xlsx":
        return {}
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        shapes: dict[str, list[int]] = {}
        for ws in wb.worksheets:
            rows = max(0, int(ws.max_row or 0) - 1)
            cols = max(0, int(ws.max_column or 0))
            shapes[str(ws.title)] = [rows, cols]
        wb.close()
        return shapes
    except Exception:
        return {}


def _normalize_sheet_name(name: str) -> str:
    value = str(name or "").strip().lower()
    value = re.sub(r"\d{4}[-_/年]?\d{1,2}[-_/月]?\d{0,2}日?", "{date}", value)
    value = re.sub(r"\d+", "{num}", value)
    value = re.sub(r"\s+", "", value)
    return value or "{sheet}"


def _header_signature(columns: list[str]) -> str:
    normalized = [re.sub(r"\s+", "", str(c or "").strip().lower()) for c in columns]
    return json.dumps(normalized, ensure_ascii=False, separators=(",", ":"))


def _header_similarity(sig_a: str, sig_b: str) -> float:
    try:
        a = set(json.loads(sig_a))
        b = set(json.loads(sig_b))
    except Exception:
        return 0.0
    if not a and not b:
        return 1.0
    return len(a & b) / max(1, len(a | b))


def _group_excel_sheet_inventories(inventories: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: list[dict[str, Any]] = []
    for inv in inventories:
        if inv.get("error"):
            continue
        sig = str(inv.get("header_signature", "") or "")
        pattern = str(inv.get("sheet_name_pattern", "") or "{sheet}")
        sheet = str(inv.get("sheet_name", "") or "")
        placed = False
        for group in groups:
            same_pattern = pattern == group.get("sheet_name_pattern")
            similar_header = _header_similarity(sig, str(group.get("header_signature", "") or "")) >= 0.75
            if similar_header and (same_pattern or len(group.get("sheets", [])) >= 1):
                group["sheets"].append(sheet)
                group["sheet_count"] = len(group["sheets"])
                placed = True
                break
        if not placed:
            groups.append(
                {
                    "group_id": f"sheet_group_{len(groups) + 1}",
                    "sheet_name_pattern": pattern,
                    "header_signature": sig,
                    "representative": sheet,
                    "sheets": [sheet],
                    "sheet_count": 1,
                    "columns": [str(c) for c in inv.get("columns", [])],
                }
            )
    return groups


def _numeric_stats(series: pd.Series) -> dict[str, float | int | bool]:
    if pd.api.types.is_bool_dtype(series):
        return {}
    coerced = pd.to_numeric(series, errors="coerce")
    valid = coerced.dropna()
    valid = valid[np.isfinite(valid.to_numpy())]
    if valid.empty:
        return {}
    arr = valid.to_numpy()
    return {
        "mean": float(np.mean(arr)),
        "median": float(np.median(arr)),
        "std": float(np.std(arr)),
        "var": float(np.var(arr)),
        "min": float(np.min(arr)),
        "max": float(np.max(arr)),
        "has_inf": bool(np.isinf(arr).any()),
    }


def _numeric_quantiles(series: pd.Series) -> dict[str, float]:
    if pd.api.types.is_bool_dtype(series):
        return {}
    coerced = pd.to_numeric(series, errors="coerce")
    valid = coerced.dropna()
    valid = valid[np.isfinite(valid.to_numpy())]
    if valid.empty:
        return {}
    return {
        "q1": float(valid.quantile(0.25)),
        "q3": float(valid.quantile(0.75)),
        "p05": float(valid.quantile(0.05)),
        "p95": float(valid.quantile(0.95)),
    }


def _datetime_stats(series: pd.Series) -> dict[str, str]:
    # 仅对“可能具有日期语义”的列做日期解析，避免把纯数值列误判为时间戳。
    if pd.api.types.is_numeric_dtype(series):
        return {}
    if not (
        pd.api.types.is_string_dtype(series)
        or pd.api.types.is_object_dtype(series)
        or pd.api.types.is_datetime64_any_dtype(series)
    ):
        return {}
    sample = series.dropna().astype(str).head(30).tolist()
    if not sample:
        return {}
    date_like_hits = 0
    for v in sample:
        s = v.strip()
        if "-" in s or "/" in s or ":" in s or "年" in s or "月" in s or "日" in s:
            date_like_hits += 1
    if date_like_hits < max(2, int(len(sample) * 0.35)):
        return {}
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        dt = pd.to_datetime(series, errors="coerce")
    dt = dt.dropna()
    if dt.empty:
        return {}
    sorted_dt = dt.sort_values()
    granularity = ""
    if len(sorted_dt) >= 2:
        diffs = sorted_dt.diff().dropna()
        if not diffs.empty:
            median_seconds = float(diffs.dt.total_seconds().median())
            if median_seconds <= 1:
                granularity = "second_or_finer"
            elif median_seconds <= 60:
                granularity = "minute"
            elif median_seconds <= 3600:
                granularity = "hour"
            elif median_seconds <= 86400:
                granularity = "day"
            elif median_seconds <= 86400 * 31:
                granularity = "month_like"
            else:
                granularity = "coarse"
    return {
        "min": dt.min().isoformat(),
        "max": dt.max().isoformat(),
        "range_days": str((dt.max() - dt.min()).days),
        "granularity": granularity,
    }


def _datetime_parse_ratio(series: pd.Series) -> float:
    if pd.api.types.is_numeric_dtype(series):
        return 0.0
    non_null = series.dropna()
    if non_null.empty:
        return 0.0
    sample = non_null.astype(str).head(200)
    date_like = sample.map(lambda v: any(tok in str(v) for tok in ["-", "/", ":", "年", "月", "日"]))
    if float(date_like.mean()) < 0.35:
        return 0.0
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        parsed = pd.to_datetime(non_null, errors="coerce")
    return round(float(parsed.notna().mean()), 6)


def _format_hints(series: pd.Series) -> list[str]:
    hints: list[str] = []
    if pd.api.types.is_integer_dtype(series):
        hints.append("integer_storage")
    elif pd.api.types.is_float_dtype(series):
        hints.append("float_storage")
    elif pd.api.types.is_bool_dtype(series):
        hints.append("boolean_storage")
    elif pd.api.types.is_datetime64_any_dtype(series):
        hints.append("datetime_storage")
    elif pd.api.types.is_string_dtype(series) or pd.api.types.is_object_dtype(series):
        hints.append("text_storage")

    vals = [str(x).strip() for x in series.dropna().astype(str).head(200).tolist() if str(x).strip()]
    if not vals:
        return hints

    def _ratio(pred) -> float:
        return sum(1 for v in vals if pred(v)) / max(1, len(vals))

    if _ratio(lambda v: bool(__import__("re").fullmatch(r"[-+]?\d+", v))) >= 0.8:
        hints.append("integer_string")
    if _ratio(lambda v: bool(__import__("re").fullmatch(r"[-+]?(?:\d+\.\d+|\d+|\.\d+)", v))) >= 0.8:
        hints.append("numeric_string")
    if _ratio(lambda v: "%" in v) >= 0.5:
        hints.append("percentage_string")
    if _ratio(lambda v: "," in v and any(ch.isdigit() for ch in v)) >= 0.5:
        hints.append("comma_number_string")
    if _ratio(lambda v: any(tok in v for tok in ["-", "/", ":", "年", "月", "日"])) >= 0.35:
        hints.append("date_time_like_string")
    if _ratio(lambda v: v.lower() in {"true", "false", "yes", "no", "y", "n", "0", "1"}) >= 0.8:
        hints.append("boolean_like_string")
    return list(dict.fromkeys(hints))


def _logical_type(series: pd.Series, numeric_ratio: float, datetime_ratio: float, unique_count: int) -> str:
    if pd.api.types.is_bool_dtype(series):
        return "boolean"
    if pd.api.types.is_datetime64_any_dtype(series) or datetime_ratio >= 0.8:
        return "datetime"
    if pd.api.types.is_integer_dtype(series):
        return "integer"
    if pd.api.types.is_float_dtype(series):
        return "float"
    if numeric_ratio >= 0.95:
        return "numeric_string"
    if numeric_ratio >= 0.6:
        return "mixed_numeric_text"
    if unique_count <= 20:
        return "categorical"
    return "text"


def column_profile_to_dict(profile: ColumnProfile) -> dict[str, Any]:
    return {
        "name": profile.name,
        "dtype": profile.dtype,
        "logical_type": profile.logical_type,
        "row_count": profile.row_count,
        "null_count": profile.null_count,
        "non_null_count": profile.non_null_count,
        "null_ratio": profile.null_ratio,
        "unique_count": profile.unique_count,
        "numeric_parse_ratio": profile.numeric_parse_ratio,
        "datetime_parse_ratio": profile.datetime_parse_ratio,
        "format_hints": profile.format_hints,
        "sample_values": profile.sample_values,
        "top_values": profile.top_values,
        "value_pattern_hints": profile.value_pattern_hints,
        "numeric_stats": profile.numeric_stats,
        "quantiles": profile.quantiles,
        "datetime_stats": profile.datetime_stats,
        "abnormal_tokens": profile.abnormal_tokens,
    }


def _value_pattern_hints(series: pd.Series) -> list[str]:
    vals = [str(x).strip() for x in series.dropna().astype(str).head(120).tolist() if str(x).strip()]
    if not vals:
        return []
    hints: list[str] = []

    def _ratio(pred) -> float:
        hit = sum(1 for v in vals if pred(v))
        return hit / max(1, len(vals))

    if _ratio(lambda v: bool(pd.notna(v)) and bool(__import__("re").match(r"^\d+\s*型[\u4e00-\u9fffA-Za-z]+$", v))) >= 0.5:
        hints.append("numbered_type_enum")
    # 脱敏编号：如“粤B****8”或“*A23**”
    if _ratio(lambda v: ("*" in v or "×" in v or "X" in v) and any(ch.isdigit() for ch in v)) >= 0.4:
        hints.append("masked_plate_like")
    # 代码标识：大写字母+数字+连接符
    if _ratio(lambda v: bool(__import__("re").match(r"^[A-Za-z0-9_-]{6,}$", v))) >= 0.6:
        hints.append("code_like")
    # 行政区文本
    if _ratio(lambda v: any(k in v for k in ["省", "市", "区", "县", "镇", "街道"])) >= 0.4:
        hints.append("region_name_like")
    return hints


def profile_dataframe(df: pd.DataFrame, top_k: int = 12) -> list[ColumnProfile]:
    def _safe_unique_count(series: pd.Series) -> int:
        try:
            return int(series.nunique(dropna=True))
        except TypeError:
            # list/dict 等不可哈希对象退化为 JSON 字符串后再统计唯一值。
            def _norm(v: Any) -> Any:
                if isinstance(v, (list, dict)):
                    try:
                        return json.dumps(v, ensure_ascii=False, sort_keys=True)
                    except TypeError:
                        return str(v)
                return v

            normalized = series.map(_norm)
            return int(normalized.nunique(dropna=True))

    profiles: list[ColumnProfile] = []
    for col in df.columns:
        s = df[col]
        row_count = int(s.shape[0])
        null_count = int(s.isna().sum())
        non_null_count = int(row_count - null_count)
        null_ratio = float(s.isna().mean())
        unique_count = _safe_unique_count(s)
        sample_values = s.dropna().astype(str).head(top_k).tolist()
        numeric_parse_ratio = 0.0
        if non_null_count > 0 and not pd.api.types.is_bool_dtype(s):
            numeric_parse_ratio = round(float(pd.to_numeric(s.dropna(), errors="coerce").notna().mean()), 6)
        datetime_parse_ratio = _datetime_parse_ratio(s)
        profile = ColumnProfile(
            name=str(col),
            dtype=str(s.dtype),
            null_ratio=null_ratio,
            unique_count=unique_count,
            row_count=row_count,
            null_count=null_count,
            non_null_count=non_null_count,
            numeric_parse_ratio=numeric_parse_ratio,
            datetime_parse_ratio=datetime_parse_ratio,
            format_hints=_format_hints(s),
            sample_values=sample_values,
        )
        profile.numeric_stats = _numeric_stats(s)
        profile.quantiles = _numeric_quantiles(s)
        profile.datetime_stats = _datetime_stats(s)
        profile.logical_type = _logical_type(s, numeric_parse_ratio, datetime_parse_ratio, unique_count)
        try:
            vc = s.dropna().astype(str).value_counts().head(top_k)
            profile.top_values = [f"{idx}({int(cnt)})" for idx, cnt in vc.items()]
        except Exception:
            profile.top_values = []
        profile.value_pattern_hints = _value_pattern_hints(s)
        # 检查“主要是数字但掺杂字符串”的情况。
        if not pd.api.types.is_bool_dtype(s):
            coerced = pd.to_numeric(s, errors="coerce")
            numeric_ratio = float(coerced.notna().mean())
            if numeric_ratio > 0.6:
                mask = coerced.isna() & s.notna()
                bad_tokens = s[mask].astype(str).value_counts().head(top_k).index.tolist()
                profile.abnormal_tokens = [str(x) for x in bad_tokens]
        profiles.append(profile)
    return profiles


def dataframe_summary(df: pd.DataFrame) -> dict[str, Any]:
    return {
        "shape": [int(df.shape[0]), int(df.shape[1])],
        "columns": [str(c) for c in df.columns.tolist()],
    }
